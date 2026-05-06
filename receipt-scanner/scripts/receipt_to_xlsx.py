"""
receipt_to_xlsx.py

영수증에서 추출한 품목 데이터를 xlsx 파일로 저장하거나 기존 파일에 누적.

사용법:
    python receipt_to_xlsx.py --input items.json --output receipts.xlsx
    python receipt_to_xlsx.py --input items.json --output receipts.xlsx --append

JSON 입력 포맷:
    [
      {
        "날짜": "2026-05-06",
        "매장명": "GS25 강남역점",
        "품목": "삼각김밥 참치마요",
        "수량": 2,
        "단가": 1500,
        "합계": 3000,
        "카테고리": "식료품"
      },
      ...
    ]

왜 JSON 입력? Claude가 영수증을 읽고 구조화한 결과를 그대로 파일로 떨궈서
이 스크립트로 넘기는 게 가장 안정적이라서. 직접 인자로 다 받으면 인자 길이/이스케이프 지옥.
"""

import argparse
import json
import os
import shutil
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 필요합니다. 설치: pip install openpyxl --break-system-packages")
    raise

SHEET_NAME = "Receipts"
COLUMNS = ["날짜", "매장명", "품목", "수량", "단가", "합계", "카테고리"]


def load_items(json_path: str) -> list[dict]:
    """JSON 파일에서 품목 리스트 읽기."""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("JSON은 품목 객체의 리스트여야 합니다.")
    return data


def validate_item(item: dict) -> dict:
    """품목 dict 검증 + 누락 필드 채우기. 잘못된 데이터는 안전한 기본값으로."""
    cleaned = {}
    for col in COLUMNS:
        cleaned[col] = item.get(col, "")
    # 숫자 필드 강제 변환
    for k in ("수량", "단가", "합계"):
        v = cleaned[k]
        if isinstance(v, str):
            v = v.replace(",", "").replace("원", "").replace("₩", "").strip()
        try:
            cleaned[k] = int(float(v)) if v != "" else 0
        except (ValueError, TypeError):
            cleaned[k] = 0
    return cleaned


def create_new_workbook() -> Workbook:
    """새 워크북 + 헤더 셋업."""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    write_header(ws)
    return wb


def write_header(ws) -> None:
    """헤더 행 굵게 + 배경색."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for idx, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def append_items_to_sheet(ws, items: list[dict]) -> None:
    """기존 시트 마지막 행 다음부터 품목 추가."""
    start_row = ws.max_row + 1 if ws.max_row > 1 else 2
    for i, item in enumerate(items):
        row = start_row + i
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            value = item.get(col_name, "")
            cell = ws.cell(row=row, column=col_idx, value=value)
            # 금액 컬럼은 우측 정렬 + 천단위 콤마
            if col_name in ("단가", "합계"):
                cell.number_format = '#,##0"원"'
                cell.alignment = Alignment(horizontal="right")
            elif col_name == "수량":
                cell.alignment = Alignment(horizontal="center")
            elif col_name == "카테고리":
                cell.alignment = Alignment(horizontal="center")


def sort_sheet_by_date(ws) -> None:
    """날짜 컬럼 기준 오름차순 정렬. 헤더 보존."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(c is not None and c != "" for c in row):
            rows.append(row)
    # 날짜 파싱 실패시 맨 뒤로
    def date_key(r):
        try:
            return datetime.strptime(str(r[0]), "%Y-%m-%d")
        except (ValueError, TypeError):
            return datetime.max
    rows.sort(key=date_key)
    # 기존 데이터 클리어
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
            # 포맷은 유지하기 위해 cell 자체는 두고 값만 비움
    # 다시 쓰기
    for i, row in enumerate(rows, start=2):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=value)
            col_name = COLUMNS[j - 1]
            if col_name in ("단가", "합계"):
                cell.number_format = '#,##0"원"'
                cell.alignment = Alignment(horizontal="right")
            elif col_name in ("수량", "카테고리"):
                cell.alignment = Alignment(horizontal="center")


def adjust_column_widths(ws) -> None:
    """컬럼 너비 자동 조정 (대략)."""
    widths = {"날짜": 12, "매장명": 22, "품목": 30, "수량": 8, "단가": 12, "합계": 14, "카테고리": 12}
    for idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col_name, 15)


def backup_corrupted_file(path: str) -> str:
    """기존 파일이 손상됐거나 컬럼이 다를 때 백업본 생성."""
    p = Path(path)
    backup_name = f"{p.stem}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{p.suffix}"
    backup_path = p.parent / backup_name
    shutil.copy2(path, backup_path)
    return str(backup_path)


def is_valid_existing_workbook(path: str) -> bool:
    """기존 xlsx의 시트명/헤더가 우리 포맷과 맞는지 검증."""
    try:
        wb = load_workbook(path)
        if SHEET_NAME not in wb.sheetnames:
            return False
        ws = wb[SHEET_NAME]
        header = [ws.cell(row=1, column=i + 1).value for i in range(len(COLUMNS))]
        return header == COLUMNS
    except Exception:
        return False


def main():
    parser = argparse.ArgumentParser(description="영수증 품목 → xlsx")
    parser.add_argument("--input", required=True, help="품목 JSON 파일 경로")
    parser.add_argument("--output", required=True, help="출력 xlsx 경로")
    parser.add_argument(
        "--append",
        action="store_true",
        help="기존 파일에 누적 (지정 안 하면 파일 존재 시 누적, 없으면 새로 생성)",
    )
    args = parser.parse_args()

    items = load_items(args.input)
    items = [validate_item(it) for it in items]

    output_exists = os.path.exists(args.output)

    if output_exists:
        if not is_valid_existing_workbook(args.output):
            backup = backup_corrupted_file(args.output)
            print(f"[경고] 기존 파일 포맷이 달라서 백업 생성: {backup}")
            print("[경고] 새 파일로 덮어쓰지 않았습니다. 사용자 확인 후 처리해주세요.")
            return 1
        wb = load_workbook(args.output)
        ws = wb[SHEET_NAME]
        append_items_to_sheet(ws, items)
    else:
        wb = create_new_workbook()
        ws = wb[SHEET_NAME]
        append_items_to_sheet(ws, items)

    sort_sheet_by_date(ws)
    adjust_column_widths(ws)
    wb.save(args.output)

    total = sum(it["합계"] for it in items)
    print(f"[성공] {len(items)}개 품목 저장 → {args.output}")
    print(f"[성공] 이번 추가분 총액: {total:,}원")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
